import customtkinter as ctk
import os
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import workbook

#Configurando a aparencia do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.apparence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Gerenciamento de Clientes")
        self.geometry("700x500")

    def apparence(self):
        self.lb_apm = ctk.CTkLabel(self, text = "Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x= 50, y= 430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command = self.change_app).place(x=50, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color = "#1E90FF", fg_color = "#1E90FF").place(x= 0, y=10)
        title = ctk.CTkLabel(frame, text= "Sistema de Gerenciamento de Clientes", font = ("Century Gothic bold", 24), 
                             text_color="white", fg_color="#1E90FF").place(x=140, y=10)
        span = ctk.CTkLabel(frame, text= "Por favor, preencha todos os campos!", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"]).place(x=50, y=70)
        
        def submit():
            #Pegando as informações dos entrys
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            adress = adress_value.get()
            gender= gender_comobox.get()
            obs = obs_entry.get(0.0, END)

            if not os.path.isfile("Clientes.xlsx"):
                # Se não existir, crie o arquivo com um cabeçalho
                arquivo = openpyxl.Workbook()
                folha = arquivo.active
                folha.append(["Nome", "Contato", "Idade", "Endereço", "Observação"])
                arquivo.save("Clientes.xlsx")

            arquivo = openpyxl.load_workbook("Clientes.xlsx")
            folha = arquivo.active
            folha.cell(column = 1, row = folha.max_row+1, value = name)
            folha.cell(column = 2, row = folha.max_row, value = contact)
            folha.cell(column = 3, row = folha.max_row, value = age)
            folha.cell(column = 4, row = folha.max_row, value = adress)
            folha.cell(column = 5, row = folha.max_row, value = obs)

            arquivo.save("Clientes.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso!")
 

        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            adress_value.set("")
            obs_entry.delete(0,0, END)

        #Variaves de texto
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        adress_value = StringVar()
        
        #Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font = ("Century Gothic bold", 16), fg_color="transparent")
        contato_entry = ctk.CTkEntry(self, width=200,textvariable=contact_value, font = ("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font = ("Century Gothic bold", 16), fg_color="transparent")
        adress_entry = ctk.CTkEntry(self, width=200, textvariable=adress_value, font = ("Century Gothic bold", 16), fg_color="transparent")

        #Combobox
        gender_comobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font= ("Century Gothic bold", 14))
        gender_comobox.set("-")
        gender_comobox.place(x=500, y=220)


        #Entrada de observações
        obs_entry = ctk.CTkTextbox(self, width= 500, height=150, font = ("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")


        #Labels
        lb_name= ctk.CTkLabel(frame, text= "Nome completo", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(frame, text= "Contato", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(frame, text= "Idade", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(frame, text= "Gênero", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        lb_adress = ctk.CTkLabel(frame, text= "Endereço", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(frame, text= "Observação", font = ("Century Gothic bold", 16), 
                             text_color=["#000", "#fff"])
        
        btn_submit = ctk.CTkButton(self, text= "Salvar".upper(), command= submit, fg_color="#32CD32", 
                                   hover_color="#3CB371",).place(x=300, y=420)

        btn_submit = ctk.CTkButton(self, text= "Limpar campos".upper(), command= clear, fg_color="#FF6347", 
                                   hover_color="#B22222").place(x=500, y=420)


        #Posicionando os elementos
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x= 450, y=120)
        contato_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y= 220)

        lb_gender.place(x=500, y=190)
        gender_comobox.place(x=500, y=220)

        lb_adress.place(x=50, y=190)
        adress_entry.place(x=50, y= 220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y= 260)


    def change_app(self, new_apparence_mode):
        ctk.set_appearance_mode(new_apparence_mode)

if __name__=="__main__":
    app = App()
    app.mainloop()